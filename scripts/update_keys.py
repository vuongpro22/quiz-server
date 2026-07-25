import os
import shutil
import re
import openpyxl
import sys

# Configure stdout for UTF-8 to prevent console encoding issues
sys.stdout.reconfigure(encoding='utf-8')

XLSX_PATH = r"d:\tool linh tinh\toolMultipleChoice\quiz-web\quiz-server\KEYS.xlsx"
QUESTION_DIR = r"D:\tool linh tinh\toolMultipleChoice\quiz-web\quiz-server\question"
BACKUP_DIR = r"D:\tool linh tinh\toolMultipleChoice\quiz-web\quiz-server\question\.backup_questions"

MAPPING = {
    "mln122-sp26-b5-fe-re.461": "MLN122 - SP26 - B5 - FE - RE.txt",
    "mln122-sp26-b5-fe.371": "MLN122 - SP26 - B5 - FE.txt",
    "mln122-SP26_C1_FE": "MLN122 - SP26 - C1 FE.txt",
    "mln122-SP26_C2_FE": "MLN122 - SP26 - C2 FE.txt",
    "mln122-SP26_RE_FE": "MLN122 - SP26 - FE - RE.txt",
    "PRM393 - SP26 - FE": "PRM393 - SP26 - FE.txt",
    "PRM393 - SP26 - B5 - FE": "PRM393 - SP26 - B5 - FE.txt",
}

def normalize_ans(val):
    if val is None:
        return ""
    val_str = str(val).strip()
    if not val_str:
        return ""
    # Extract only alphabetical chars
    letters = re.findall(r'[A-Za-z]', val_str)
    # Uppercase
    letters = [l.upper() for l in letters]
    # Sort
    letters.sort()
    # Join with comma and space
    return ", ".join(letters)

def find_file_path(filename):
    for root, dirs, files in os.walk(QUESTION_DIR):
        if ".backup_questions" in root:
            continue
        if filename in files:
            return os.path.join(root, filename)
    return None

def main():
    print("Starting update keys script...")
    
    # 1. Ensure backup directory exists
    if not os.path.exists(BACKUP_DIR):
        os.makedirs(BACKUP_DIR)
        print(f"Created backup directory: {BACKUP_DIR}")
        
    # 2. Open Excel workbook
    wb = openpyxl.load_workbook(XLSX_PATH)
    sheet = wb.active
    
    # Read headers
    headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    
    # Collect all rows
    all_rows = list(sheet.iter_rows(min_row=3, values_only=True))
    
    # Let's iterate over each header
    for col_idx, h in enumerate(headers):
        if h not in MAPPING:
            print(f"Skipping header not in mapping: {h}")
            continue
            
        filename = MAPPING[h]
        filepath = find_file_path(filename)
        
        if not filepath:
            print(f"Error: Could not find file path for {filename}")
            continue
            
        # Extract all answers for this column
        answers = []
        for r in all_rows:
            answers.append(r[col_idx])
            
        # Find the last non-empty answer to know the total question count
        last_non_empty_idx = -1
        for idx in range(len(answers) - 1, -1, -1):
            if answers[idx] is not None and str(answers[idx]).strip() != "":
                last_non_empty_idx = idx
                break
        
        num_questions = last_non_empty_idx + 1
        normalized_answers = [normalize_ans(answers[i]) for i in range(num_questions)]
        
        print(f"\nProcessing {filename} (Mapped to '{h}'):")
        print(f"  Total questions detected in Excel: {num_questions}")
        
        # 3. Read original file
        with open(filepath, "r", encoding="utf-8") as f:
            content = f.read()
            
        lines = content.splitlines()
        
        # Find the header line "Câu\tĐáp án"
        header_idx = -1
        for idx, line in enumerate(lines):
            if "Câu" in line and "Đáp án" in line:
                header_idx = idx
                break
                
        if header_idx == -1:
            print(f"  Error: 'Câu\\tĐáp án' header not found in {filename}. Skipping.")
            continue
            
        # Keep everything up to and including the header line
        new_lines = lines[:header_idx+1]
        
        # Add new answer rows
        for q_num, ans in enumerate(normalized_answers, 1):
            new_lines.append(f"{q_num}\t{ans}")
            
        # Join and add trailing newline
        new_content = "\n".join(new_lines) + "\n"
        
        # 4. Backup the file
        backup_path = os.path.join(BACKUP_DIR, filename)
        shutil.copy2(filepath, backup_path)
        print(f"  Backed up original to: {backup_path}")
        
        # 5. Write the updated content
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(new_content)
        print(f"  Successfully updated {filename} with new answers.")
        
    print("\nAll done!")

if __name__ == "__main__":
    main()
